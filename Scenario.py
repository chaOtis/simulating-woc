from Interval import *
from Event import *
from Expert import *
import openpyxl
import numpy as np
from Model import Model
from Algorithm import *
from ResultData import *
import sys

class Scenario():

    def __init__(self, id):
        self.name = id
        self.readCSV(id)


    def __init__(self, excel=False, excelName="./Scenarios.xlsx", id=None, cycles=None, target=None, seeds=None, numIntervals=None, experts=None):
        self.name = id
        self.excelName = excelName
        if(excel == True):
            self.readCSV(id)
        else:
            self.cycles = cycles
            self.target = target
            self.seeds = seeds
            self.experts = experts
            self.numIntervals = numIntervals
            self.intervals = getIntervalsForEvent(numIntervals, self.target)

    def copy(self):
        return Scenario(excel=False, id=self.name, cycles=self.cycles, target=self.target,
                        seeds=self.seeds, numIntervals=self.numIntervals, experts=self.experts)


    def readCSV(self, id):
        wb = openpyxl.load_workbook(self.excelName)
        sheet = wb.get_sheet_by_name(id)
        numExperts = sheet["B4"].value
        numIntervals = sheet["B5"].value
        self.numIntervals = numIntervals
        numSeeds = sheet["B6"].value
        numEpsilons = sheet["B7"].value
        self.cycles = sheet["B8"].value

        epsilon_mu = []
        epsilon_sigma = []
        for i in range(5, 5+numEpsilons):
            epsilon_mu.append(sheet.cell(row=7, column=i).value)
            epsilon_sigma.append(sheet.cell(row=8, column=i).value)

        expert_delta = []
        expert_sigma = []
        for i in range(11, 11+numExperts):
            expert_delta.append(sheet["A"+str(i)].value)
            expert_sigma.append(sheet["B"+str(i)].value)

        sigma = np.power(epsilon_sigma,2)
        nsquared = np.power(len(epsilon_sigma),2)
        sigma = np.sqrt((sum(sigma))/nsquared)

        epsilon_mu = list(np.asarray(epsilon_mu)*4*sigma)
        expert_delta = list(np.asarray(expert_delta)*4*sigma)
        expert_sigma = list(np.asarray(expert_sigma)*sigma)

        self.target = Event(epsilon_mu, epsilon_sigma)
        self.seeds = [Event(epsilon_mu, epsilon_sigma) for i in range(0,numSeeds)]
        self.intervals = getIntervalsForEvent(numIntervals, self.target)
        self.experts = []
        for i in range(11, 11+numExperts):
            expert_epsilon = []
            for j in range(5, 5+numEpsilons):
                expert_epsilon.append(sheet.cell(row=i, column=j).value)
            self.experts.append(Expert(expert_delta[i-11], expert_sigma[i-11], expert_epsilon))


    def simulate(self, x=0, n=0):
        rps_scores = [["uwm","pwm","cwm","con","bex","ran"]]
        p_performance = []
        c_performance = []
        p_weights = []
        c_weights = []
        n_weights = []
        b_weights = []
        self.intervals = getIntervalsForEvent(self.numIntervals, self.target)

        for i in range(0, self.cycles):
            rps = []
            self.target.setObservation()
            for s in self.seeds:
                s.setObservation()
            sim = Model(self.experts, self.intervals, self.seeds, self.target)
            data = sim.getJudgmentData()

            uwm = UWM(data)
            pwm = PWM(data)
            cwm = CWM(data)
            con = Contribution(data)
            bex = BestExpert(data)
            ran = RandomExpert(data)

            rps.append(uwm.run())
            rps.append(pwm.run())
            rps.append(cwm.run())
            con.setPM(cwm.getPM())
            rps.append(con.run())
            bex.setPM(pwm.getPM())
            rps.append(bex.run())
            rps.append(ran.run())
            rps_scores.append(rps)

            p_performance.append(pwm.getPM())
            c_performance.append(cwm.getPM())
            p_weights.append(pwm.getWeights())
            c_weights.append(cwm.getWeights())
            n_weights.append(con.getWeights())
            b_weights.append(bex.getWeights())
            sys.stderr.write('\r')
            # the exact output you're looking for:
            sys.stderr.write("["+str(int(x))+"/"+str(n)+"] "+"[%-100s] %d%%" % ('=' * int((i+1)/self.cycles*100), (i+1)/self.cycles*100))
            sys.stderr.flush()
        result = ResultData(rps_scores, self.name)
        result.setExpertPerformance([p_performance,c_performance])
        result.setExpertWeights([p_weights, c_weights, n_weights, b_weights])
        sys.stderr.write("\n")
        return result


    def simulateDefaultWeights(self, weights):
        rps_scores = [["dew"]]
        self.intervals = getIntervalsForEvent(self.numIntervals, self.target)
        for i in range(0, self.cycles):
            rps = []
            self.target.setObservation()
            for s in self.seeds:
                s.setObservation()
            sim = Model(self.experts, self.intervals, self.seeds, self.target)
            data = sim.getJudgmentData()

            dew = DefaultWeight(data)
            dew.setWeights(weights)
            rps.append(dew.run())
            rps_scores.append(rps)
            sys.stderr.write('\r')
            # the exact output you're looking for:
            sys.stderr.write("[%-100s] %d%%" % ('=' * int((i+1)/self.cycles*100), (i+1)/self.cycles*100))
            sys.stderr.flush()
        result = ResultData(rps_scores, self.name)
        sys.stderr.write("\n")
        return result


    def writeCSV(self):
        pass